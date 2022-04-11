import openpyxl


def read_reactor_from_xlsx(folder, filename):
    """
    cascade = {
        reactor: {
            'Наименование': str,\n
            'Температура вход': float,\n
            'Температура выход': float,\n
            'Давление вход': float,\n
            'Давление выход': float,\n
            'Объем': float,\n
            'Кол-во шагов': float,\n
            'Мольный расход': float,\n
        }
    }

    :return: All reactor from xlsx file
    :rtype: dict
    """

    file = f'{folder}/{filename}'

    xlsx = openpyxl.load_workbook(file, data_only=True)
    sheet = xlsx.active

    cascade = {}
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(column=1, row=row).value
        cascade[name] = {}
        for col in range(2, sheet.max_column + 1):
            try:
                cascade[name][sheet.cell(column=col, row=1).value] = float(sheet.cell(column=col, row=row).value)
            except ValueError:
                cascade[name][sheet.cell(column=col, row=1).value] = sheet.cell(column=col, row=row).value
    return cascade